import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.db.models import Q

from apps.catalogos.models import NumeroParte
from apps.conceptos.models import (
    Concepto,
    DocumentoConceptos,
    HistorialCoincidencia,
    PatronSerie,
    construir_firma_concepto,
)


COLUMNAS_ESPERADAS = (
    'numero_parte',
    'serie',
    'modelo',
    'descripcion',
    'cantidad',
    'precio_unitario',
)


def analizar_archivo_conceptos(archivo, documento=None):
    filas_crudas = _leer_archivo(archivo)
    series_archivo = set()
    filas = [
        _analizar_fila(indice, fila, documento=documento, series_archivo=series_archivo)
        for indice, fila in enumerate(filas_crudas, start=2)
    ]
    resumen = {
        'filas_totales': len(filas),
        'validas': sum(1 for fila in filas if fila['valida']),
        'con_sugerencia': sum(1 for fila in filas if fila['sugerida']),
        'incompletas': sum(1 for fila in filas if fila['estado'] == 'incompleto'),
        'errores': sum(1 for fila in filas if fila['estado'] == 'error'),
        'duplicadas': sum(1 for fila in filas if fila['estado'] == 'duplicada'),
    }
    return {'resumen': resumen, 'filas': filas}


def confirmar_importacion_conceptos(
    documento,
    filas,
    usuario=None,
    usar_para_biblioteca=False,
):
    creados = 0
    orden_base = documento.conceptos.count()
    for fila in filas:
        if not fila.get('valida'):
            continue
        datos = fila['datos']
        concepto = Concepto.objects.create(
            documento=documento,
            numero_parte=datos['numero_parte'],
            serie=datos['serie'],
            modelo=datos['modelo'],
            descripcion=datos['descripcion'],
            cantidad=Decimal(datos['cantidad']),
            precio_unitario=Decimal(datos['precio_unitario']),
            orden=orden_base + creados + 1,
        )
        if usar_para_biblioteca:
            _registrar_historial_importacion(concepto, fila, usuario)
        creados += 1
    documento.recalcular_total()
    return creados


def _registrar_historial_importacion(concepto, fila, usuario=None):
    firma_texto, firma_json = construir_firma_concepto(concepto)
    HistorialCoincidencia.objects.get_or_create(
        concepto=concepto,
        defaults={
            'documento': concepto.documento,
            'serie': concepto.serie,
            'numero_parte': concepto.numero_parte,
            'modelo': concepto.modelo,
            'descripcion': concepto.descripcion,
            'firma_texto': firma_texto,
            'firma_json': firma_json,
            'regla_usada': 'importacion_confirmada',
            'match_type': _match_type_importacion(fila),
            'confirmado_por': usuario if getattr(usuario, 'is_authenticated', False) else None,
            'confirmado_en_importacion': True,
            'usar_para_biblioteca': True,
        },
    )


def _match_type_importacion(fila):
    return {
        'ok_exacto': 'exacto',
        'sugerido_historial': 'historial',
        'sugerido_patron': 'patron',
        'manual': 'manual',
    }.get(fila.get('estado'), 'manual')


def _leer_archivo(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith('.csv'):
        texto = archivo.read().decode('utf-8-sig')
        return _leer_csv(texto)
    if nombre.endswith('.xlsx'):
        return _leer_xlsx(archivo)
    raise ValueError('Formato no soportado. Usa CSV o XLSX.')


def _leer_csv(texto):
    reader = csv.DictReader(io.StringIO(texto))
    reader.fieldnames = [_normalizar_header(campo) for campo in (reader.fieldnames or [])]
    return [
        {_normalizar_header(clave): valor for clave, valor in fila.items()}
        for fila in reader
    ]


def _leer_xlsx(archivo):
    from openpyxl import load_workbook

    archivo.seek(0)
    workbook = load_workbook(archivo, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalizar_header(valor) for valor in rows[0]]
    filas = []
    for row in rows[1:]:
        filas.append({
            header: '' if valor is None else str(valor)
            for header, valor in zip(headers, row)
        })
    return filas


def _normalizar_header(valor):
    valor = str(valor or '').strip().lower().replace('\ufeff', '')
    return re.sub(r'\s+', '_', valor)


def _analizar_fila(numero_fila, fila, documento=None, series_archivo=None):
    datos = {
        columna: str(fila.get(columna) or '').strip()
        for columna in COLUMNAS_ESPERADAS
    }
    series_archivo = series_archivo if series_archivo is not None else set()
    serie_normalizada = _normalizar_serie(datos['serie'])
    cantidad, error_cantidad = _decimal_o_default(datos['cantidad'], Decimal('1'))
    precio, error_precio = _decimal_o_default(datos['precio_unitario'], Decimal('0'))
    datos['cantidad'] = str(cantidad) if cantidad is not None else datos['cantidad']
    datos['precio_unitario'] = str(precio) if precio is not None else datos['precio_unitario']
    errores = []

    if serie_normalizada and _serie_existe_en_documento(documento, serie_normalizada):
        estado = 'duplicada'
        mensaje = 'La serie ya existe en este documento.'
    elif serie_normalizada and serie_normalizada in series_archivo:
        estado = 'duplicada'
        mensaje = 'La serie está repetida dentro del archivo.'
    else:
        if serie_normalizada:
            series_archivo.add(serie_normalizada)
        estado, mensaje = _validar_y_sugerir_fila(datos, cantidad, precio, error_cantidad, error_precio)

    total = Decimal('0') if cantidad is None or precio is None else cantidad * precio
    return {
        'fila': numero_fila,
        'datos': datos,
        'total_concepto': str(total),
        'estado': estado,
        'estado_label': _estado_label(estado),
        'mensaje': mensaje,
        'valida': estado not in ('incompleto', 'error', 'duplicada'),
        'sugerida': estado in ('ok_exacto', 'sugerido_historial', 'sugerido_patron'),
    }


def _validar_y_sugerir_fila(datos, cantidad, precio, error_cantidad, error_precio):
    errores = []
    if not any([datos['numero_parte'], datos['serie'], datos['descripcion']]):
        errores.append('Debe capturar numero_parte, serie o descripcion.')
        estado = 'incompleto'
    else:
        estado = 'manual'

    if error_cantidad or cantidad is None or cantidad <= 0:
        errores.append('Cantidad invalida.')
        estado = 'error'
    if error_precio or precio is None or precio < 0:
        errores.append('Precio unitario invalido.')
        estado = 'error'

    if not errores:
        estado, mensaje = _aplicar_sugerencia(datos)
    else:
        mensaje = ' '.join(errores)

    return estado, mensaje


def _normalizar_serie(serie):
    return (serie or '').strip().upper()


def _serie_existe_en_documento(documento, serie_normalizada):
    if not documento or not serie_normalizada:
        return False
    series = documento.conceptos.exclude(serie='').values_list('serie', flat=True)
    return any(_normalizar_serie(serie) == serie_normalizada for serie in series)


def _decimal_o_default(valor, default):
    if valor == '':
        return default, False
    try:
        return Decimal(str(valor)), False
    except (InvalidOperation, TypeError, ValueError):
        return None, True


def _aplicar_sugerencia(datos):
    parte = _buscar_numero_parte_activo(datos['numero_parte'])
    if parte:
        if not datos['modelo']:
            datos['modelo'] = parte.modelo
        if not datos['descripcion']:
            datos['descripcion'] = parte.descripcion
        return 'ok_exacto', 'NumeroParte activo encontrado.'

    concepto = _buscar_historial_confirmado(datos)
    if concepto:
        if not datos['modelo']:
            datos['modelo'] = concepto.modelo
        if not datos['descripcion']:
            datos['descripcion'] = concepto.descripcion
        if not datos['numero_parte']:
            datos['numero_parte'] = concepto.numero_parte
        return 'sugerido_historial', 'Sugerido por historial confirmado.'

    patron = _buscar_patron_serie(datos['serie'])
    if patron:
        if not datos['modelo']:
            datos['modelo'] = patron.modelo
        if not datos['descripcion']:
            datos['descripcion'] = patron.descripcion
        if not datos['numero_parte']:
            datos['numero_parte'] = patron.numero_parte
        return 'sugerido_patron', 'Sugerido por patron de serie.'

    return 'manual', 'Fila valida sin sugerencia.'


def _buscar_numero_parte_activo(numero_parte):
    if not numero_parte:
        return None
    try:
        parte = NumeroParte.objects.get(numero_parte=numero_parte.strip())
    except NumeroParte.DoesNotExist:
        return None
    return parte if parte.activo else None


def _buscar_historial_confirmado(datos):
    criterios = Q()
    if datos['numero_parte']:
        criterios |= Q(numero_parte__iexact=datos['numero_parte'])
    if datos['serie']:
        criterios |= Q(serie__iexact=datos['serie'])
    if datos['descripcion']:
        criterios |= Q(descripcion__icontains=datos['descripcion'])
    if not criterios:
        return None
    return (
        Concepto.objects.filter(documento__status=DocumentoConceptos.STATUS_CONFIRMADO)
        .exclude(numero_parte='', serie='', descripcion='')
        .filter(criterios)
        .order_by('-updated_at', '-id')
        .first()
    )


def _buscar_patron_serie(serie):
    serie = (serie or '').strip().upper()
    if not serie:
        return None
    patrones = PatronSerie.objects.filter(
        activo=True,
        campo_identificador=PatronSerie.CAMPO_SERIE,
    )
    matches = [patron for patron in patrones if serie.startswith(patron.prefix)]
    if not matches:
        return None
    return sorted(matches, key=lambda item: (-len(item.prefix), item.prefix))[0]


def _estado_label(estado):
    labels = {
        'ok_exacto': 'OK exacto',
        'sugerido_historial': 'Sugerido por historial',
        'sugerido_patron': 'Sugerido por patron',
        'incompleto': 'Incompleto',
        'error': 'Error',
        'duplicada': 'Duplicada',
        'manual': 'Manual',
    }
    return labels.get(estado, estado)
