import csv
import re
import unicodedata
from datetime import datetime
from io import StringIO

from apps.catalogos.models import ClaveProductoServicioSAT, NumeroParte

MAX_FILAS_NUMEROS_PARTE = 5000
MAX_MUESTRA_PREVIEW = 10

_ENCABEZADOS_NUMEROS_PARTE = {
    'numero_parte': {'numero_parte', 'numero_de_parte', 'numero_parte_catalogo'},
    'modelo': {'modelo'},
    'descripcion': {'descripcion'},
    'fraccion': {'fraccion'},
}


def importar_numeros_parte_csv(archivo):
    preview = analizar_numeros_parte_csv(archivo)
    resultado = guardar_numeros_parte_desde_filas(preview['filas'])
    resultado['errores'] = preview['errores']
    return resultado


def analizar_numeros_parte_csv(archivo, limite=MAX_FILAS_NUMEROS_PARTE):
    preview = {
        'archivo_nombre': getattr(archivo, 'name', ''),
        'filas_validas': 0,
        'crearian': 0,
        'actualizarian': 0,
        'errores': [],
        'muestra': [],
        'filas': [],
    }
    numeros_en_archivo = set()
    limite_excedido = False
    encabezados = None

    for fila_numero, fila in _leer_archivo_numeros_parte(archivo):
        if encabezados is None:
            if not any(_normalizar_texto(valor) for valor in fila):
                continue
            encabezados = _mapear_encabezados_numero_parte(fila)
            faltantes = [
                nombre for nombre in ('numero_parte', 'descripcion')
                if nombre not in encabezados
            ]
            if faltantes:
                _agregar_error(
                    preview,
                    fila_numero,
                    'Faltan encabezados obligatorios: ' + ', '.join(faltantes) + '.',
                )
                break
            continue

        valores = _normalizar_fila_por_encabezados(fila, encabezados)
        numero_parte = valores['numero_parte']
        modelo = valores['modelo']
        descripcion = valores['descripcion']
        fraccion = valores['fraccion']
        if not any(valores.values()):
            continue

        if len(preview['filas']) + len(preview['errores']) >= limite:
            limite_excedido = True
            break

        if not numero_parte:
            _agregar_error(preview, fila_numero, 'numero_parte es requerido.')
            continue
        if not descripcion:
            _agregar_error(preview, fila_numero, 'descripcion es requerida.')
            continue
        if numero_parte in numeros_en_archivo:
            _agregar_error(preview, fila_numero, 'numero_parte duplicado dentro del archivo.')
            continue

        numeros_en_archivo.add(numero_parte)
        fila_normalizada = {
            'fila': fila_numero,
            'numero_parte': numero_parte,
            'modelo': modelo,
            'descripcion': descripcion,
            'fraccion': fraccion,
        }
        preview['filas'].append(fila_normalizada)
        if len(preview['muestra']) < MAX_MUESTRA_PREVIEW:
            preview['muestra'].append(fila_normalizada)

    if limite_excedido:
        _agregar_error(preview, limite + 1, f'El archivo excede el limite de {limite} filas.')

    existentes = set()
    if preview['filas']:
        existentes = set(
            NumeroParte.objects.filter(
                numero_parte__in=[fila['numero_parte'] for fila in preview['filas']]
            ).values_list('numero_parte', flat=True)
        )
    preview['filas_validas'] = len(preview['filas'])
    preview['total_filas'] = preview['filas_validas'] + len(preview['errores'])
    preview['actualizarian'] = sum(
        1 for fila in preview['filas'] if fila['numero_parte'] in existentes
    )
    preview['crearian'] = preview['filas_validas'] - preview['actualizarian']
    return preview


def guardar_numeros_parte_desde_filas(filas):
    resultado = _resultado_base()

    for fila in filas:
        _, creado = NumeroParte.objects.update_or_create(
            numero_parte=fila['numero_parte'],
            defaults={
                'modelo': fila['modelo'],
                'descripcion': fila['descripcion'],
                'fraccion': fila['fraccion'],
            },
        )
        resultado['procesadas'] += 1
        _contar_guardado(resultado, creado)

    return resultado


_ENCABEZADOS_CARTA_PORTE = {
    'clave': {'c_claveprodserv', 'cclaveprodserv', 'clave'},
    'descripcion': {'descripcion'},
    'palabras_similares': {'palabras_similares', 'palabrassimilares'},
    'material_peligroso': {'material_peligroso', 'materialpeligroso'},
    'fecha_inicio_vigencia': {'fechainiciovigencia', 'fecha_inicio_vigencia'},
    'fecha_fin_vigencia': {'fechafinvigencia', 'fecha_fin_vigencia'},
    'incluir_iva_trasladado': {'incluir_iva_trasladado'},
    'incluir_ieps_trasladado': {'incluir_ieps_trasladado'},
    'complemento_que_debe_incluir': {'complemento_que_debe_incluir'},
    'estimulo_franja_fronteriza': {'estimulo_franja_fronteriza'},
}


def importar_claves_sat_csv(archivo):
    resultado = _resultado_base()
    filas = _leer_filas_carta_porte(archivo)

    for fila_numero, fila in filas:
        if not any(str(valor or '').strip() for valor in fila.values()):
            continue

        resultado['procesadas'] += 1

        clave = _normalizar_clave_carta_porte(fila.get('clave'))
        if not clave:
            _agregar_error(resultado, fila_numero, 'clave es requerida.')
            continue
        if not _parece_clave_sat(clave):
            _agregar_error(resultado, fila_numero, 'clave debe ser numerica de 8 digitos.')
            continue

        descripcion = str(fila.get('descripcion') or '').strip()
        if not descripcion:
            _agregar_error(resultado, fila_numero, 'descripcion es requerida.')
            continue

        _, creado = ClaveProductoServicioSAT.objects.update_or_create(
            clave=clave,
            defaults={
                'descripcion': descripcion,
                'incluir_iva_trasladado': fila.get('incluir_iva_trasladado', ''),
                'incluir_ieps_trasladado': fila.get('incluir_ieps_trasladado', ''),
                'complemento_que_debe_incluir': fila.get('material_peligroso') or fila.get('complemento_que_debe_incluir', ''),
                'fecha_inicio_vigencia': _parsear_fecha(fila.get('fecha_inicio_vigencia')),
                'fecha_fin_vigencia': _parsear_fecha(fila.get('fecha_fin_vigencia')),
                'estimulo_franja_fronteriza': fila.get('estimulo_franja_fronteriza', ''),
                'palabras_similares': fila.get('palabras_similares', ''),
            },
        )
        _contar_guardado(resultado, creado)

    return resultado


def _leer_filas_carta_porte(archivo):
    extension = getattr(archivo, 'name', '').lower()
    if extension.endswith('.xlsx'):
        return _leer_filas_carta_porte_xlsx(archivo)

    filas = list(_leer_csv(archivo))
    encabezados = _encontrar_encabezados_carta_porte(
        (numero, fila) for numero, fila in filas
    )
    if encabezados:
        return _filas_por_encabezados(filas[encabezados[0] - 1:], encabezados[1])

    # Compatibilidad con el CSV SAT histórico sin encabezados oficiales.
    return (
        (numero, {
            'clave': valores[0] if valores else '',
            'descripcion': valores[1] if len(valores) > 1 else '',
            'incluir_iva_trasladado': valores[2] if len(valores) > 2 else '',
            'incluir_ieps_trasladado': valores[3] if len(valores) > 3 else '',
            'complemento_que_debe_incluir': valores[4] if len(valores) > 4 else '',
            'fecha_inicio_vigencia': valores[5] if len(valores) > 5 else '',
            'fecha_fin_vigencia': valores[6] if len(valores) > 6 else '',
            'estimulo_franja_fronteriza': valores[7] if len(valores) > 7 else '',
            'palabras_similares': valores[8] if len(valores) > 8 else '',
        })
        for numero, valores in filas
        if _parece_clave_sat(str(valores[0]).strip() if valores else '')
    )


def _leer_filas_carta_porte_xlsx(archivo):
    from openpyxl import load_workbook

    archivo.seek(0)
    workbook = load_workbook(archivo, read_only=True, data_only=True)
    try:
        if 'c_ClaveProdServCP' not in workbook.sheetnames:
            raise ValueError('No se encontró la hoja c_ClaveProdServCP.')
        filas = [
            (numero, list(fila))
            for numero, fila in enumerate(
                workbook['c_ClaveProdServCP'].iter_rows(values_only=True), start=1
            )
        ]
        encabezados = _encontrar_encabezados_carta_porte(iter(filas))
        if not encabezados:
            raise ValueError('No se encontraron los encabezados c_ClaveProdServ y Descripción.')
        return _filas_por_encabezados(filas[encabezados[0] - 1:], encabezados[1])
    finally:
        workbook.close()


def _encontrar_encabezados_carta_porte(filas):
    for numero, fila in filas:
        mapeo = {}
        for indice, valor in enumerate(fila):
            encabezado = _normalizar_encabezado_carta_porte(valor)
            for nombre, alias in _ENCABEZADOS_CARTA_PORTE.items():
                if encabezado in alias and nombre not in mapeo:
                    mapeo[nombre] = indice
                    break
        if 'clave' in mapeo and 'descripcion' in mapeo:
            return numero, mapeo
    return None


def _filas_por_encabezados(filas, encabezados):
    for numero, fila in filas[1:]:
        yield numero, {
            nombre: _valor_fila(fila, indice)
            for nombre, indice in encabezados.items()
        }


def _valor_fila(fila, indice):
    if indice >= len(fila) or fila[indice] is None:
        return ''
    return fila[indice]


def _normalizar_encabezado_carta_porte(valor):
    texto = _sin_acentos(str(valor or '')).strip().lower()
    return re.sub(r'[^a-z0-9]+', '_', texto).strip('_')


def _normalizar_clave_carta_porte(valor):
    if valor is None:
        return ''
    texto = str(valor).strip()
    if texto.endswith('.0') and texto[:-2].isdigit():
        texto = texto[:-2]
    return texto.zfill(8) if texto.isdigit() else texto


def _resultado_base():
    return {
        'procesadas': 0,
        'creadas': 0,
        'actualizadas': 0,
        'errores': [],
    }


def _leer_csv(archivo):
    archivo.seek(0)
    contenido = archivo.read()
    if isinstance(contenido, bytes):
        texto = _decodificar(contenido)
    else:
        texto = contenido

    lector = csv.reader(StringIO(texto))
    for indice, fila in enumerate(lector, start=1):
        yield indice, fila


def _leer_archivo_numeros_parte(archivo):
    extension = getattr(archivo, 'name', '').lower()
    if extension.endswith('.csv'):
        yield from _leer_csv(archivo)
        return
    if extension.endswith('.xlsx'):
        yield from _leer_xlsx(archivo)
        return
    raise ValueError('Formato no soportado. Usa CSV o XLSX.')


def _leer_xlsx(archivo):
    from openpyxl import load_workbook

    archivo.seek(0)
    workbook = load_workbook(archivo, read_only=True, data_only=True)
    try:
        for indice, fila in enumerate(
            workbook.active.iter_rows(values_only=True), start=1
        ):
            yield indice, [valor if valor is not None else '' for valor in fila]
    finally:
        workbook.close()


def _decodificar(contenido):
    for codificacion in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            return contenido.decode(codificacion)
        except UnicodeDecodeError:
            continue
    return contenido.decode('utf-8', errors='replace')


def _normalizar_fila(fila, longitud):
    valores = [str(valor).strip() for valor in fila[:longitud]]
    return valores + [''] * (longitud - len(valores))


def _normalizar_texto(valor):
    texto = _sin_acentos(str(valor or '')).strip().lower()
    return re.sub(r'[^a-z0-9]+', '_', texto).strip('_')


def _mapear_encabezados_numero_parte(fila):
    encabezados = {}
    for indice, valor in enumerate(fila):
        normalizado = _normalizar_texto(valor)
        for nombre, alias in _ENCABEZADOS_NUMEROS_PARTE.items():
            if normalizado in alias and nombre not in encabezados:
                encabezados[nombre] = indice
                break
    return encabezados


def _normalizar_fila_por_encabezados(fila, encabezados):
    return {
        nombre: str(fila[indice]).strip() if indice < len(fila) and fila[indice] is not None else ''
        for nombre, indice in (
            ('numero_parte', encabezados.get('numero_parte')),
            ('modelo', encabezados.get('modelo')),
            ('descripcion', encabezados.get('descripcion')),
            ('fraccion', encabezados.get('fraccion')),
        )
        if indice is not None
    } | {
        nombre: ''
        for nombre in ('numero_parte', 'modelo', 'descripcion', 'fraccion')
        if nombre not in encabezados
    }


def _parece_encabezado_numero_parte(valor):
    normalizado = _sin_acentos(valor).lower()
    return 'numero' in normalizado or 'parte' in normalizado


def _sin_acentos(valor):
    return ''.join(
        caracter
        for caracter in unicodedata.normalize('NFKD', valor)
        if not unicodedata.combining(caracter)
    )


def _parece_clave_sat(valor):
    return bool(re.fullmatch(r'\d{8}', valor or ''))


def _parsear_fecha(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, 'year') and hasattr(valor, 'month') and hasattr(valor, 'day'):
        return valor

    valor = str(valor).strip()

    for formato in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(valor, formato).date()
        except ValueError:
            continue
    return None


def _agregar_error(resultado, fila, error):
    resultado['errores'].append({'fila': fila, 'error': error})


def _contar_guardado(resultado, creado):
    if creado:
        resultado['creadas'] += 1
    else:
        resultado['actualizadas'] += 1
